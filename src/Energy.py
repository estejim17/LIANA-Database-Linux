#importacion de selenium y modulos relacionados
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

import requests
import progressbar
import string
import datetime

import numpy as np
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import os
import sys
import time
import re
from openpyxl import load_workbook

import ExtractmirRNA

import atexit

def exit_handler():
    if safe == 0:
        df1 = output
        try:
            del df1['Energy Tot']
            del df1['Index Match']
        except:
            print('Not necessary to delete columns in dataframe...')
        df2 = pd.DataFrame()
        df2['Energy Tot'] = energies
        df2['Index Match'] = matches
        df1 = pd.concat([df1,df2], ignore_index=True, axis=1)
        df1.columns = ['lncRNA','Cadena','CONTRA-FOLD','C-Energía','VIENNA','V-Energía',
                        'query','ref','chromosome','m_start','m_end','energy','score','conserve', 'cancer related', 'Energy Tot', 'Index Match']

        print('Writing to Excel file...')
        df1.to_excel(writer, j)
        writer.save()
        print('Done! Exited the program and saved unexpectedly.')
atexit.register(exit_handler)

#se checkea la version de chrome y si es necesario descarga el driver para selenium

#se extraen miRNA usando la funcion de ExtractmirRNA.py

# worksheets = pd.ExcelFile('/home/estejim15/LIANA-Database-Linux/database/output.xlsx')
# listmiRNA = worksheets.sheet_names
# listmiRNA.insert(0,listmiRNA.pop())
listmiRNA = ExtractmirRNA.getmiRNA()

# wb = load_workbook('./database/output.xlsx')

# keep_sheets = ['hsa-let-7a-5p']
# for sheetName in wb.sheetnames:
#     if sheetName not in keep_sheets:
#         # del wb[sheetName]
#         try:
#             wb.remove(wb[sheetName])
#             print('DELETED' + sheetName)
#         except:
#             print('could not delete ' + sheetName)
# wb.save('./database/output.xlsx')

def remove_space(string):
    return "".join(string.rstrip().lstrip())

def dataRead(sheet1):
    #se leen los archivos excel para extraer los datos necesarios.
    #se necesitan: las cadenas para buscar en linearfold, los demas datos ya extraidos
    #de linearfold para continuar con la extraccion
    output = pd.read_excel('./database/output.xlsx', sheet1, index_col=0, engine='openpyxl')
    sequence = output.values[:,1].tolist()
    structure = output.values[:,2].tolist()
    ref = output.values[:,7].tolist()
    if "Energy Tot" not in output:
        energies = []
        matches = []
        inicio = 0
    else:
        if len(output['Energy Tot'].dropna()) == 0:
            energies = []
            matches = []
            inicio = 0        
        else:
            energies = output['Energy Tot'].dropna().tolist()
            matches = output['Index Match'].dropna().tolist()
            inicio = len(energies)
        

    
    return output, sequence, structure, energies, ref, inicio, matches

link = "http://rna.tbi.univie.ac.at//cgi-bin/RNAWebSuite/RNAeval.cgi"


#se crea instancia de selenium
caps = DesiredCapabilities().CHROME
caps["pageLoadStrategy"] = "eager"  #  para que no espere a que se cargue la pagina
driver = webdriver.Chrome(executable_path=r'/usr/local/bin/chromedriver',  desired_capabilities=caps)
action = ActionChains(driver)


    
counter = 1
for j in listmiRNA:
    safe = 0
    #se recorren todos los miRNA (cada worksheet de excel es uno)
    output, sequence, structure, energies, ref, inicio, matches = dataRead(j)
    

    print(str(counter) + "- "+ j)  
    counter += 1



    if inicio != len(sequence): 
        #para evitar que se lean todos los excel aunque ya hayan
        #sido completados
    
        #se borran los worksheets en los que se esta trabajando
        #para luego crear un duplicado y evitar que hayan dos iguales
        # if inicio < len(sequence):
        #     wb = load_workbook('./database/output.xlsx') 
        #     wb.remove(wb[j])
        #     wb.save('./database/output.xlsx')
    
        #se crea instancia de ExcelWriter para leer el excel 
        options = {}
        options['strings_to_formulas'] = False
        options['strings_to_urls'] = False
        #pylint: disable=abstract-class-instantiated
        writer = pd.ExcelWriter('./database/output.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace')
        for i in range(inicio, len(sequence)): 

            #se accede al sitio web con selenium y se sigue el workflow necesario
            #en la pagina
            driver.get(link)
            while(len(driver.find_elements_by_id("SEQUENCE"))<=0 ):   
                driver.get(link)       

            try:            
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID,'SEQUENCE' ))
                )
                
                if sequence[i] != 'ERROR':
                    driver.find_element_by_id("SEQUENCE").send_keys(Keys.CONTROL, 'a')
                    driver.find_element_by_id("SEQUENCE").send_keys(sequence[i])
                    time.sleep(0.5)
                    driver.find_element_by_id("STRUCTURE").send_keys(Keys.CONTROL, 'a')
                    driver.find_element_by_id("STRUCTURE").send_keys(structure[i])
                    time.sleep(0.5)                        
                    driver.find_element_by_xpath('//*[@id="contentmain"]/form/div[4]/table/tbody/tr/td[2]/input').click()
                else:
                    break
            except:
                fecha = str(datetime.datetime.now())
                print("Page refreshed at time: " + fecha + " because it didn't load properly")
                time.sleep(1)
           
            try:
            
                    element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH,'//*[@id="contentmain"]/div[1]/textarea' ))
                    )
                    description = driver.find_element_by_xpath('//*[@id="contentmain"]/div[1]/textarea').get_attribute('innerHTML')
                    energies.append(description[description.rfind("(")+1:description.rfind("\n")-1])
                    # print(energies)
                    # print(sequence[i])
                    # print(ref[i])
                    to_find = remove_space(ref[i]).lower().replace('-','')
                    index_start = sequence[i].lower().find(to_find)
                    idx = str(index_start) + '-' + str(index_start+len(to_find))
                    matches.append(idx)
                   
                    
                
                    print('Copiando secuencia {}'.format(i))
                    
            except:
                print('Fallo en cargar datos.')
                # linearfoldC.append('ERROR')
                energies.append('ERROR')
                matches.append('ERROR')
                # linearfoldV.append('ERROR')
                # energiaV.append('ERROR')

        #se crean distintos dataframes para unir todos los datos al final de la extraccion
        #luego se escriben en el excel y se guarda
        df1 = output
        try:
            del df1['Energy Tot']
            del df1['Index Match']
        except:
            print('Not necessary to delete columns in dataframe...')
        df2 = pd.DataFrame()
        df2['Energy Tot'] = energies
        df2['Index Match'] = matches
        df1 = pd.concat([df1,df2], ignore_index=True, axis=1)
        df1.columns = ['lncRNA','Cadena','CONTRA-FOLD','C-Energía','VIENNA','V-Energía',
                        'query','ref','chromosome','m_start','m_end','energy','score','conserve', 'cancer related', 'Energy Tot', 'Index Match']

        print('Writing to Excel file...')
        df1.to_excel(writer, j)
        writer.save()
    safe = 1

print('Done!')
